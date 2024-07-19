# Below are notes to self, ignore when debugging
# dbo.tbl_cms_Roles
# Right click -> Select top 1000 rows

import pyodbc
import openpyxl
import os
import db_connection

def all(value):

    conn = db_connection.connect()

    # Table name
    tbl_name = value

    # Execute SQL query
    cursor = conn.cursor()
    query = f"SELECT * FROM {tbl_name}"
    cursor.execute(query)

    # Get column names
    column_names = [column[0] for column in cursor.description]

    # Fetch all rows
    rows = cursor.fetchall()

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write column names to the first row
    for col_num, column_name in enumerate(column_names, start=1):
        worksheet.cell(row=1, column=col_num, value=column_name)

    # Write data to the Excel worksheet starting from the second row
    for row_num, row in enumerate(rows, start=2):
        for col_num, value in enumerate(row, start=1):
            worksheet.cell(row=row_num, column=col_num, value=value)

    # Auto-fit column widths based on the longest value in each column
    for col_num, column_data in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value)) for cell in column_data)
        column_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet.column_dimensions[column_letter].width = max_length + 2

    # Get the path to the Downloads folder
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")

    # Construct the full file path and full file path
    output_file_name = tbl_name + ".xlsx"
    output_file = os.path.join(downloads_path, output_file_name)

    # Save the Excel file in the Downloads folder
    workbook.save(output_file)
    print(f"Table named [{tbl_name}] saved in {downloads_path}")

    # Close the database connection
    cursor.close()
    conn.close()